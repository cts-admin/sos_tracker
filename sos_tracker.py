"""
sos_tracker.py
==============
App to handle gps coordinate upload and precipitation tracking
to aid in intelligent time management of seed collection from
native plant populations over a large area.

Author can be reached at uslaner.avery@gmail.com
"""
import csv
import datetime
import functools
import os
import random
import re
import tempfile
import timeit
import xml.etree.ElementTree as ET

from flask import (Flask, flash, g, redirect, render_template, request, Response,
	send_from_directory, session, url_for)
from flask_bcrypt import check_password_hash
from flask_googlemaps import GoogleMaps, Map
from flask_login import LoginManager, login_user, logout_user, login_required, current_user
from openpyxl import Workbook
from peewee import *
from playhouse.flask_utils import get_object_or_404, object_list
from urllib.parse import urlencode
from werkzeug.utils import secure_filename

import models
import forms

app = Flask(__name__)
app.config.from_object(__name__)
app.config.from_envvar('SOS_TRACKER_SETTINGS')

GoogleMaps(app)

login_manager = LoginManager()
login_manager.init_app(app)
login_manager.long_view = 'login'

@login_manager.user_loader
def load_user(userid):
	try:
		return models.User.get(models.User.id == userid)
	except models.DoesNotExist:
		return None

@app.before_request
def before_request():
	"""Connect to the database before each request."""
	g.db = models.DATABASE
	g.db.connect()
	g.user = current_user

@app.after_request
def after_request(response):
	"""Close the database connection after each request."""
	g.db.close()
	return response

@app.route('/register', methods=['GET', 'POST'])
def register():
	"""Register new users."""
	form = forms.RegisterForm()
	form.team.choices = [(team.id, team.name) for team in models.Team.select().order_by('name')]
	if form.validate_on_submit():
		flash("Yay, you registered!", "success")
		models.User.create_user(
			username = form.username.data,
			email = form.email.data,
			password = form.password.data,
			team = form.team.data
		)
		user = models.User.get(models.User.email == form.email.data)
		# Log in the newly registered user
		login_user(user)
		return redirect(url_for('index'))
	return render_template('register.html', form=form)

@app.route('/login', methods=['GET', 'POST'])
def login():
	"""Log in existing users."""
	form = forms.LoginForm()
	if form.validate_on_submit():
		try:
			user = models.User.get(models.User.email == form.email.data)
		except:
			flash("Your email or password doesn't match!", "error")
		else:
			if check_password_hash(user.password, form.password.data):
				login_user(user)
				flash("You've been logged in!", "success")
				return redirect(url_for('index'))
			else:
				flash("Your email or password doesn't match!", "error")
	return render_template('login.html', form=form)

@app.route('/logout', methods=['GET', 'POST'])
def logout():
	if request.method == 'POST':
		logout_user()
		flash("You've been logged out!", "success")
		return redirect(url_for('login'))
	return render_template('logout.html')

# Parse file into list of coordinates
def parse_file(filename, publish):
	file = os.path.join(app.config['UPLOAD_FOLDER'], filename)
	parsed = []
	user = models.User.select().where(models.User.username == g.user._get_current_object().username)
	if filename[-3:] == 'txt':
		with open(file, 'r') as f:
			csv_input = csv.reader(f)

			for line in csv_input:
				# To avoid index out of bound errors
				if len(line) >= 4:
					slug = re.sub('[^\w]+', '-', line[2].lower())
					if any(point['slug'] == slug for point in parsed):
						# The idea here is to create a 1 in 1000 chance of collision for each subsequent non-unique
						# name the slugs are being generated from
						slug = slug + str(random.randint(0, 999))
					coordinate = {'user': user,
						'latitude': line[0],
						'longitude': line[1],
						'name': line[2],
						'pin': line[3],
						'notes': 'Uploaded from file.',
						'published': publish,
						'slug': slug
					}
					parsed.append(coordinate)

	elif filename[-3:] == 'gpx':
		tree = ET.parse(file)
		root = tree.getroot()
		for child in root:
			if child.tag[-3:] == 'wpt':
				lat = child.attrib['lat']
				lon = child.attrib['lon']
				name = None
				pin = None
				for tag in child:
					if tag.tag[-4:] == 'name':
						name = tag.text
					elif tag.tag[-3:] == 'sym':
						pin = tag.text

				if name and pin:
					slug = re.sub('[^\w]+', '-', name.lower())
					if any(point['slug'] == slug for point in parsed):
						# The idea here is to create a 1 in 1000 chance of collision for each subsequent non-unique
						# name the slugs are being generated from
						slug = slug + str(random.randint(0, 999))
					coordinate = {'user': user,
						'latitude': lat,
						'longitude': lon,
						'name': name,
						'pin': pin,
						'notes': 'Uploaded from file.',
						'published': publish,
						'slug': slug
					}
					parsed.append(coordinate)
	return parsed

# Database expects parsed data
# data is a list of coordinate dictionaries
def save_to_database(data):
	error = 0
	with models.DATABASE.atomic():
		try:
			models.Coordinate.insert_many(data).execute()
		except IntegrityError as e:
			error += 1
			print(e.args)
	if error > 0:
		flash('{} point(s) were not able to be added. Point names must be unique.'.format(error), 'danger')
	# error = 0
	# for coord in data:
	# 	try:
	# 		point = models.Coordinate.create(
	# 			user = g.user._get_current_object(),
	# 			latitude = coord['Latitude'],
	# 			longitude = coord['Longitude'],
	# 			name = coord['Name'],
	# 			pin = coord['Pin'],
	# 			notes = 'Uploaded from file.',
	# 			published = publish
	# 		)
	# 	except IntegrityError as e:
	# 		error += 1
	# 		print(e.args)
	# if error > 0:
	# 	flash('{} points were not able to be added. Point names must be unique.'.format(error), 'danger')

def write_workbook(query, filename, search):
	"""Create ArcGIS compatible xls file using openpyxl."""
	wb = Workbook()
	ws = wb.active
	ws['A1'] = 'Latitude'
	ws['B1'] = 'Longitude'
	ws['C1'] = 'Name'
	ws['D1'] = 'Pin'
	ws['E1'] = 'Notes'
	for idx, point in enumerate(query, 2):
		if search:
			point = point.point
		ws['A' + str(idx)] = point.latitude
		ws['B' + str(idx)] = point.longitude
		ws['C' + str(idx)] = point.name
		ws['D' + str(idx)] = point.pin
		ws['E' + str(idx)] = point.notes

	return wb

# Views
# Main landing page
@app.route('/')
def index():
	search_query = request.args.get('q')
	if search_query:
		query = models.Coordinate.search(search_query)
	else:
		query = models.Coordinate.public().order_by(models.Coordinate.timestamp.desc())
	return object_list('index.html', query, search=search_query, check_bounds=False)

# Manually create a single GPS point
@app.route('/create', methods = ['GET', 'POST'])
@login_required
def create():
	form = forms.CreateCoordForm()
	if form.validate_on_submit():
		user = models.User.select().where(models.User.username == g.user._get_current_object().username)
		point = models.Coordinate.create(
			user = user,
			latitude = form.latitude.data,
			longitude = form.longitude.data,
			name = form.name.data,
			pin = form.pin.data,
			notes = form.notes.data,
			published = form.published.data
		)
		flash("Entry created successfully.", "success")			
		return redirect(url_for('detail', slug=point.slug))
	return render_template('create.html', form=form)

# Download points from database to ArcGIS compatible .xls file format
@app.route('/download', methods = ['GET', 'POST'])
@login_required
def download():
	search_query = request.args.get('dq') or request.form.get('prev-search')
	if search_query:
		query = models.Coordinate.search(search_query)
		search = True
	else:
		query = models.Coordinate.public().order_by(models.Coordinate.timestamp.desc())
		search = False
	if request.method == 'POST':
		if request.form.get('filename'):
			filename = secure_filename(request.form.get('filename')) + '.xls'
			wb = write_workbook(query, filename, search)
			if request.form.get('save'):
				wb.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))				
				flash('Search added to files successfully!', 'success')
				return send_from_directory(app.config['UPLOAD_FOLDER'], filename, as_attachment=True)
			else:
				temp = tempfile.TemporaryDirectory()
				wb.save(temp.name + '/' + filename)
				return send_from_directory(temp.name, filename, as_attachment=True)
		else:
			flash('Please specify a search name.', 'danger')
	return object_list('download.html', query, dsearch=search_query, check_bounds=False)

# View unpublished points
@app.route('/private')
@login_required
def private():
	query = models.Coordinate.private().order_by(models.Coordinate.timestamp.desc())
	return object_list('index.html', query, check_bounds=False)

@app.route('/<slug>')
def detail(slug):
	if current_user.is_authenticated:
		query = models.Coordinate.select()
	else:
		query = models.Coordinate.public()
	point = get_object_or_404(query, models.Coordinate.slug == slug)
	if point:
		pointmap = Map(
			identifier="pointmap",
			varname="pointmap",
			lat=point.latitude,
			lng=point.longitude,
			markers=[(point.latitude, point.longitude)]
		)
	return render_template('detail.html', point=point, pointmap=pointmap)

@app.route('/<slug>/edit', methods=['GET', 'POST'])
@login_required
def edit(slug):
	point = get_object_or_404(models.Coordinate, models.Coordinate.slug == slug)
	if request.method == 'POST':
		if request.form.get('latitude') and request.form.get('longitude') and request.form.get('name'):
			point.latitude = request.form['latitude']
			point.longitude = request.form['longitude']
			point.name = request.form['name']
			point.pin = request.form['pin']
			point.notes = request.form['notes']
			point.published = request.form.get('published') or False
			point.save()

			flash('Point saved successfully!', 'success')
			if point.published:
				return redirect(url_for('detail', slug=point.slug))
			else:
				return redirect(url_for('edit', slug=point.slug))
		else:
			flash("'Latitude', 'Longitude', and 'Name', are required fields.", 'danger')

	return render_template('edit.html', point=point)

# From peewee blog example
# http://charlesleifer.com/blog/how-to-make-a-flask-blog-in-one-hour-or-less/
@app.template_filter('clean_querystring')
def clean_querystring(request_args, *keys_to_remove, **new_values):
	querystring = dict((key, value) for key, value in request_args.items())
	for key in keys_to_remove:
		querystring.pop(key, None)
	querystring.update(new_values)
	return urlencode(querystring)

# View list of uploaded files
@app.route('/files')
def list_files():
	files = os.listdir(app.config['UPLOAD_FOLDER'])
	return render_template('files.html', files=files)

@app.errorhandler(404)
def not_found(exc):
	return Response('<h3>Not Found!</h3>'), 404

# Upload files
@app.route('/upload', methods=['GET', 'POST'])
def upload():
	form = forms.CoordFile()
	if form.validate_on_submit():
		# Save uploaded file on server if it is valid
		filename = secure_filename(form.file.data.filename)	
		publish = form.publish.data
		form.file.data.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
		data = parse_file(filename, publish)
		save_to_database(data)
		flash('File uploaded successfully!', 'success')
		return redirect(url_for('index'))
	return render_template('upload.html', form=form)

# View file
@app.route('/upload/<filename>')
def uploaded(filename):
	return send_from_directory(app.config['UPLOAD_FOLDER'], filename)

def main():
	models.initialize()
	try:
		models.Team.create_team(
			name='Red Butte Garden',
			institution='University of Utah',
			code='RBG'
		)
	except ValueError:
		pass
	app.run(debug=True)

if __name__ == '__main__':
	main()